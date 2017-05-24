#!usr/bin/python3
"""
Block builder used in TSEA83 project. Takes an excel document and 
translates standard 6-digit hex color codes to 2-digit codes.
USAGE: python(3) block_builder.py [FILE] [START CELL] [END CELL]
"""

import sys
import openpyxl
from openpyxl import load_workbook

xlsx_file = sys.argv[1]
start = sys.argv[2]
end = sys.argv[3]
with open("blocks", "w+") as fd:
    ws = load_workbook(filename=xlsx_file).active
    i = 0
    cell_range = ws[start:end]
    for col in cell_range:
        for cell in col:
            i += 1
            color = cell.fill.start_color.index

            red = int(color[2:4], 16)
            green = int(color[4:6], 16)
            blue = int(color[6:8], 16)
            red_8 = (red * 7) // 255
            green_8 = (green * 7) // 255
            blue_8 = (blue * 3) // 255

            binary_val = format(red_8, '03b') + format(green_8, '03b') + format(blue_8, '02b')
            output = format(int(binary_val, 2), "02x")

            fd.write("x\"" + output + "\", ")
            if i % 300 == 0:
                fd.write("\n")
        fd.write("\n")
